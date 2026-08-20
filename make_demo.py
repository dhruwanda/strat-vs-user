"""Generate the fictional demo dataset.

The story (kept deliberately simple, all weights 25% always):
  Jan 15  first investment, Rs 1,00,000 across Albus, Bellatrix, Cedric, Draco
  Feb     rebalance: Draco out, Dobby in       (investor applies 2 days late)
  Mar     rebalance: Albus out, Aragog in      (2 days late)
  Mar 10  invest-more, another Rs 1,00,000
  Apr     rebalance: Bellatrix out, Buckbeak in (2 days late)

Prices drift mostly upward so the demo reads as a healthy portfolio; the
dropped names are the laggards. The investor's 2-day lags and intraday fills
make the model-vs-you lines visibly diverge. One dividend belongs to a stock
the smallcase never held (NIMBUS) so the exclusion logic has something to show.
Entirely fictional; no real data anywhere.
"""
import numpy as np
import pandas as pd
import openpyxl

rng = np.random.default_rng(11)
OUT = "demo_data"

SYMS = ["ALBUS", "BELLATRIX", "CEDRIC", "DRACO", "DOBBY", "ARAGOG", "BUCKBEAK"]
NAMES = {"ALBUS": "Albus Industries Ltd", "BELLATRIX": "Bellatrix Power Ltd",
         "CEDRIC": "Cedric Pharma Ltd", "DRACO": "Draco Metals Ltd",
         "DOBBY": "Dobby Logistics Ltd", "ARAGOG": "Aragog Textiles Ltd",
         "BUCKBEAK": "Buckbeak Aviation Ltd"}
P0 = {"ALBUS": 520.0, "BELLATRIX": 145.0, "CEDRIC": 880.0, "DRACO": 62.0,
      "DOBBY": 240.0, "ARAGOG": 410.0, "BUCKBEAK": 96.0}
# mostly positive drift; the names that get dropped are the laggards
DRIFT = {"ALBUS": -.0002, "BELLATRIX": .0002, "CEDRIC": .0016, "DRACO": -.0018,
         "DOBBY": .0018, "ARAGOG": .0022, "BUCKBEAK": .0014}

cal = pd.bdate_range("2026-01-02", "2026-06-30")
px = {}
for s in SYMS:
    steps = rng.normal(DRIFT[s], 0.011, len(cal))
    close = P0[s] * np.exp(np.cumsum(steps))
    spread = np.abs(rng.normal(0.006, 0.002, len(cal)))
    d = pd.DataFrame({"date": cal, "close": close,
                      "open": close * (1 + rng.normal(0, 0.003, len(cal)))})
    d["high"] = np.maximum(d["open"], d["close"]) * (1 + spread)
    d["low"] = np.minimum(d["open"], d["close"]) * (1 - spread)
    px[s] = d

def bar(s, d, f):
    r = px[s][px[s].date == d].iloc[0]
    return float((r.open + r.high + r.low + r.close) / 4) if f == "ohlc" else float(r[f])

def next_td(d):
    return cal[cal > d][0]

W = 0.25
V1 = ("2026-01-02 to 2026-02-01", ["ALBUS", "BELLATRIX", "CEDRIC", "DRACO"])
V2 = ("2026-02-02 to 2026-03-01", ["ALBUS", "BELLATRIX", "CEDRIC", "DOBBY"])
V3 = ("2026-03-02 to 2026-04-05", ["BELLATRIX", "CEDRIC", "DOBBY", "ARAGOG"])
V4 = ("2026-04-06 to 2026-06-30", ["CEDRIC", "DOBBY", "ARAGOG", "BUCKBEAK"])
FLAGS = [pd.Timestamp(x) for x in
         ("2026-01-02", "2026-02-02", "2026-03-02", "2026-04-06")]

# ---- official index: fractional quantities, T1 OHLC transitions ----
def transition(qold, names, t1):
    inter = sum(qq * bar(s, t1, "ohlc") for s, qq in qold.items())
    return {s: inter * W / bar(s, t1, "ohlc") for s in names}

q = {s: 100 * W / bar(s, cal[0], "close") for s in V1[1]}
t1 = {f: next_td(f) for f in FLAGS[1:]}
idx = []
for d in cal:
    for f, names in zip(FLAGS[1:], (V2[1], V3[1], V4[1])):
        if d == t1[f]:
            q = transition(q, names, d)
    idx.append(dict(date=d, val=sum(qq * bar(s, d, "close") for s, qq in q.items()),
                    flag=d in FLAGS))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Historical Index Values"
ws.append(["Date", "Index Value", "Rebalance Occured"])
for r in idx:
    ws.append([r["date"].strftime("%Y-%m-%d"), f"{r['val']:.2f}",
               "True" if r["flag"] else None])
ws2 = wb.create_sheet("Historical Constituents")
ws2.append(["Date Range", "Constituents", "Weightage"])
for rngtxt, names in (V1, V2, V3, V4):
    first = True
    for s in names:
        ws2.append([rngtxt if first else None, NAMES[s], W]); first = False
wb.save(f"{OUT}/timeline.xlsx")

# ---- investor tradebook ----
trades, tid, hold = [], [1000], {}
def add(sym, day, tstr, side, qty, price):
    tid[0] += 1
    trades.append(dict(symbol=sym, isin=f"INE{tid[0]}D01",
                       trade_date=day.strftime("%Y-%m-%d"), exchange="NSE",
                       segment="EQ", series="EQ", trade_type=side, auction=False,
                       quantity=int(qty), price=round(price, 2),
                       trade_id=tid[0], order_id=tid[0],
                       order_execution_time=f"{day.strftime('%Y-%m-%d')}T{tstr}"))

def fill(s, d, side):
    """Intraday fill: buys land a touch below mid, sells a touch above."""
    r = px[s][px[s].date == d].iloc[0]
    frac = 0.45 if side == "buy" else 0.60
    return float(r.low + frac * (r.high - r.low))

d1 = pd.Timestamp("2026-01-15"); A = 100000.0
for s in V1[1]:
    p = fill(s, d1, "buy"); qq = round(A * W / p)
    add(s, d1, "15:20:59", "buy", qq, p); hold[s] = hold.get(s, 0) + qq

def rebalance(day, names, tsell, tbuy):
    V = sum(hold[s] * fill(s, day, "sell") for s in hold if hold[s] > 0)
    tgt = {s: round(V * W / fill(s, day, "buy")) for s in names}
    for s in list(hold):
        if s not in names and hold[s] > 0:
            add(s, day, tsell, "sell", hold[s], fill(s, day, "sell")); hold[s] = 0
    for s in names:
        cur = hold.get(s, 0); dq = tgt[s] - cur
        if dq < 0:
            add(s, day, tsell, "sell", -dq, fill(s, day, "sell"))
        elif dq > 0:
            add(s, day, tbuy, "buy", dq, fill(s, day, "buy"))
        hold[s] = tgt[s]

# investor applies each rebalance 2 trading days after T1
rebalance(cal[cal > t1[FLAGS[1]]][1], V2[1], "09:30:10", "09:30:12")
rebalance(cal[cal > t1[FLAGS[2]]][1], V3[1], "09:31:00", "09:31:02")
d4 = pd.Timestamp("2026-03-10"); A2 = 100000.0
V = sum(hold[s] * fill(s, d4, "buy") for s in hold if hold[s] > 0)
for s in V3[1]:
    p = fill(s, d4, "buy")
    need = max(0.0, (V + A2) * W - hold.get(s, 0) * p)
    qq = round(need / p)
    if qq:
        add(s, d4, "09:45:11", "buy", qq, p); hold[s] = hold.get(s, 0) + qq
rebalance(cal[cal > t1[FLAGS[3]]][1], V4[1], "09:32:00", "09:32:02")

tb = pd.DataFrame(trades)
tb.to_csv(f"{OUT}/tradebook.csv", index=False)

# ---- P&L workbook ----
last = cal[-1]
tb["val"] = tb.quantity * tb.price
stt = float((tb["val"] * 0.001).sum())
stamp = float((tb.loc[tb.trade_type == "buy", "val"] * 0.00015).sum())
exch = float((tb["val"] * 0.0000297).sum())
sebi = float((tb["val"] * 0.000001).sum())
ipft = float((tb["val"] * 0.0000001).sum())
gst = 0.18 * (exch + sebi + ipft)
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Equity"
ws.append(["Client ID", "DEMO42"])
ws.append(["P&L Statement for Equity from 2026-01-01 to 2026-06-30"])
ws.append(["Summary"])
hrows, realized_total = [], 0.0
for s in SYMS:
    b = tb[(tb.symbol == s) & (tb.trade_type == "buy")]
    sl = tb[(tb.symbol == s) & (tb.trade_type == "sell")]
    if not len(b):
        continue
    bq, bv, sq, sv = b.quantity.sum(), b.val.sum(), sl.quantity.sum(), sl.val.sum()
    avg = bv / bq
    rp = sv - avg * sq
    realized_total += rp
    oq = bq - sq
    oc = avg * oq
    cp = bar(s, last, "close")
    hrows.append([s, f"INE{s[:3]}01", float(sq), round(avg * sq, 2), round(sv, 2),
                  round(rp, 2), 0.0, round(cp, 2), float(oq), "", round(oc, 2),
                  round(oq * cp - oc, 2), 0.0])
unreal_total = sum(r[11] for r in hrows)
n_dp = int((tb[tb.trade_type == "sell"].groupby(["symbol", "trade_date"]).size()).shape[0])
ws.append(["Charges", round(stt + stamp + exch + sebi + ipft + gst, 4)])
ws.append(["Other Credit & Debit", -(2 * 118.0 + n_dp * 15.0)])
ws.append(["Realized P&L", round(realized_total, 2)])
ws.append(["Unrealized P&L", round(unreal_total, 2)])
ws.append(["Charges"]); ws.append(["Account Head", "Amount"])
for k, v in [("Brokerage", 0.0), ("Exchange Transaction Charges", exch),
             ("Integrated GST", gst), ("Securities Transaction Tax", stt),
             ("SEBI Turnover Fees", sebi), ("Stamp Duty", stamp), ("IPFT", ipft)]:
    ws.append([k, round(v, 4)])
ws.append(["Symbol", "ISIN", "Quantity", "Buy Value", "Sell Value",
           "Realized P&L", "Realized P&L Pct.", "Previous Closing Price",
           "Open Quantity", "Open Quantity Type", "Open Value",
           "Unrealized P&L", "Unrealized P&L Pct."])
for r in hrows:
    ws.append(r)
ws2 = wb.create_sheet("Other Debits and Credits")
ws2.append(["Client ID", "DEMO42"])
ws2.append(["Other Debits and Credits for EQ from 2026-01-01 to 2026-06-30"])
ws2.append(["Particulars", "Posting Date", "Debit", "Credit"])
for d in (d1, d4):
    ws2.append([f"Being smallcase fee for {d.strftime('%d/%m/%Y')}",
                d.strftime("%Y-%m-%d"), 118.0, 0.0])
for (s, dd), _ in tb[tb.trade_type == "sell"].groupby(["symbol", "trade_date"]):
    day = pd.Timestamp(dd)
    ws2.append([f"DP Charges for Sale of {s} on {day.strftime('%d/%m/%Y')}",
                day.strftime("%Y-%m-%d"), 15.0, 0.0])
wb.save(f"{OUT}/pnl.xlsx")

pd.DataFrame([
    {"Symbol": "CEDRIC", "Ex-date": "2026-04-20", "Qty": int(hold["CEDRIC"]),
     "Dividend per share": 11.0, "Total dividend": 11.0 * int(hold["CEDRIC"])},
    {"Symbol": "DOBBY", "Ex-date": "2026-05-15", "Qty": int(hold["DOBBY"]),
     "Dividend per share": 3.0, "Total dividend": 3.0 * int(hold["DOBBY"])},
    {"Symbol": "NIMBUS", "Ex-date": "2026-05-20", "Qty": 40,
     "Dividend per share": 2.5, "Total dividend": 100.0},
]).to_csv(f"{OUT}/dividends.csv", index=False)

pd.concat([px[s].assign(symbol=s, volume=0)[
    ["symbol", "date", "open", "high", "low", "close", "volume"]]
    for s in SYMS]).to_csv(f"{OUT}/prices_cache.csv", index=False)
print("demo written; final holdings:", {k: int(v) for k, v in hold.items() if v})
