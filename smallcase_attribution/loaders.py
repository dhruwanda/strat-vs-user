"""
Input loaders. Each returns a normalised frame plus a provenance note so the
report can state exactly where every number came from.

Expected inputs (all optional except the tradebook and the smallcase timeline):
  * smallcase timeline workbook  - 'Historical Index Values' + 'Historical Constituents'
  * broker tradebook csv         - one row per executed trade
  * broker P&L workbook          - holdings snapshot, charge totals, other debits/credits
"""
from __future__ import annotations
import re
import pandas as pd
import numpy as np
import openpyxl


# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------
def _sheet_rows(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()
    return rows


def _find_header(rows, *required):
    """Locate the header row by the labels it must contain. Broker exports carry
    a variable number of preamble rows, so the position cannot be assumed."""
    req = {r.lower() for r in required}
    for i, r in enumerate(rows):
        cells = {str(c).strip().lower() for c in r if c is not None}
        if req <= cells:
            return i
    raise ValueError(f"could not locate a header row containing {sorted(req)}")


# ----------------------------------------------------------------------------
# smallcase timeline
# ----------------------------------------------------------------------------
def load_index_values(path: str) -> pd.DataFrame:
    """Official smallcase index series. Columns: date, index_value, rebalance_flag."""
    rows = _sheet_rows(path, "Historical Index Values")
    if rows is None:
        raise ValueError("workbook has no 'Historical Index Values' sheet")
    h = _find_header(rows, "date", "index value")
    hdr = [str(c).strip().lower() if c else "" for c in rows[h]]
    ci, cv = hdr.index("date"), hdr.index("index value")
    cr = hdr.index("rebalance occured") if "rebalance occured" in hdr else None
    rec = []
    for r in rows[h + 1:]:
        if r[ci] in (None, ""):
            continue
        flag = False
        if cr is not None and r[cr] is not None:
            flag = str(r[cr]).strip().lower() in ("true", "yes", "1")
        rec.append((pd.to_datetime(r[ci]), float(r[cv]), flag))
    df = pd.DataFrame(rec, columns=["date", "index_value", "rebalance_flag"])
    return df.sort_values("date").reset_index(drop=True)


def load_constituents(path: str) -> pd.DataFrame:
    """
    Constituent history. The 'Date Range' cell is populated only on the first row
    of each version block and must be forward-filled.
    Columns: version, version_start, version_end, name, weight.
    """
    rows = _sheet_rows(path, "Historical Constituents")
    if rows is None:
        raise ValueError("workbook has no 'Historical Constituents' sheet")
    h = _find_header(rows, "date range", "constituents", "weightage")
    hdr = [str(c).strip().lower() if c else "" for c in rows[h]]
    cd, cn, cw = hdr.index("date range"), hdr.index("constituents"), hdr.index("weightage")
    cur, rec = None, []
    for r in rows[h + 1:]:
        if r[cd] not in (None, ""):
            cur = str(r[cd]).strip()
        if r[cn] in (None, "") or cur is None:
            continue
        rec.append((cur, str(r[cn]).strip(), float(r[cw])))
    df = pd.DataFrame(rec, columns=["version", "name", "weight"])
    parts = df["version"].str.split(r"\s+to\s+", n=1, expand=True)
    df["version_start"] = pd.to_datetime(parts[0])
    df["version_end"] = pd.to_datetime(parts[1])
    return df.sort_values(["version_start", "name"]).reset_index(drop=True)


# ----------------------------------------------------------------------------
# broker tradebook
# ----------------------------------------------------------------------------
_TRADE_ALIASES = {
    "symbol": ["symbol", "tradingsymbol", "scrip", "instrument"],
    "isin": ["isin"],
    "trade_date": ["trade_date", "date", "trade date"],
    "exchange": ["exchange", "exch"],
    "trade_type": ["trade_type", "type", "buy_sell", "transaction_type"],
    "quantity": ["quantity", "qty"],
    "price": ["price", "rate", "trade_price"],
    "order_id": ["order_id", "orderid"],
    "trade_id": ["trade_id", "tradeid"],
    "order_execution_time": ["order_execution_time", "execution_time", "order_time", "time"],
}


def load_tradebook(path: str) -> pd.DataFrame:
    raw = pd.read_csv(path)
    low = {c.strip().lower(): c for c in raw.columns}
    col = {}
    for want, alts in _TRADE_ALIASES.items():
        for a in alts:
            if a in low:
                col[want] = low[a]
                break
    for req in ("symbol", "trade_date", "trade_type", "quantity", "price"):
        if req not in col:
            raise ValueError(f"tradebook is missing a '{req}' column")

    df = pd.DataFrame({k: raw[v] for k, v in col.items()})
    df["symbol"] = df["symbol"].astype(str).str.strip().str.upper()
    df["trade_type"] = df["trade_type"].astype(str).str.strip().str.lower()
    df["trade_date"] = pd.to_datetime(df["trade_date"])
    df["quantity"] = df["quantity"].astype(float)
    df["price"] = df["price"].astype(float)
    if "order_execution_time" in df:
        df["exec_ts"] = pd.to_datetime(df["order_execution_time"])
    else:
        # Without execution timestamps, basket detection falls back to trade date.
        df["exec_ts"] = df["trade_date"]
    if "exchange" not in df:
        df["exchange"] = "NSE"
    df["value"] = df["quantity"] * df["price"]
    df["signed_qty"] = np.where(df["trade_type"] == "buy", df["quantity"], -df["quantity"])
    return df.sort_values(["exec_ts", "symbol"]).reset_index(drop=True)


# ----------------------------------------------------------------------------
# broker P&L workbook
# ----------------------------------------------------------------------------
def load_pnl(path: str) -> dict:
    """
    Returns:
      holdings   - per-symbol realised/open snapshot incl. the closing price used
      charges    - account-level charge heads (NOT per trade)
      summary    - broker's own realised / unrealised / charges totals
      period     - (start, end) covered by the statement
    """
    rows = _sheet_rows(path, "Equity")
    if rows is None:
        raise ValueError("P&L workbook has no 'Equity' sheet")

    period = (None, None)
    for r in rows[:12]:
        for c in r:
            m = re.search(r"from\s+(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", str(c or ""))
            if m:
                period = (pd.to_datetime(m.group(1)), pd.to_datetime(m.group(2)))

    h = _find_header(rows, "symbol", "quantity", "buy value")
    hdr = [str(c).strip() if c is not None else "" for c in rows[h]]
    keep = [i for i, c in enumerate(hdr) if c]
    data = []
    for r in rows[h + 1:]:
        if r[keep[0]] in (None, ""):
            continue
        data.append([r[i] for i in keep])
    hold = pd.DataFrame(data, columns=[hdr[i] for i in keep])
    ren = {
        "Symbol": "symbol", "ISIN": "isin", "Quantity": "closed_qty",
        "Buy Value": "matched_buy_value", "Sell Value": "sell_value",
        "Realized P&L": "broker_realized_pnl", "Previous Closing Price": "close_price",
        "Open Quantity": "open_qty", "Open Value": "open_cost",
        "Unrealized P&L": "broker_unrealized_pnl",
    }
    hold = hold.rename(columns={k: v for k, v in ren.items() if k in hold.columns})
    hold["symbol_raw"] = hold["symbol"].astype(str).str.strip().str.upper()
    # Broker suffixes some ETF rows (e.g. '-E'); the tradebook does not.
    hold["symbol"] = hold["symbol_raw"].str.replace(r"-E$", "", regex=True)
    for c in ("closed_qty", "matched_buy_value", "sell_value", "broker_realized_pnl",
              "close_price", "open_qty", "open_cost", "broker_unrealized_pnl"):
        if c in hold:
            hold[c] = pd.to_numeric(hold[c], errors="coerce")

    # account-level charge heads sit above the holdings table
    charges, summary = {}, {}
    for r in rows[:h]:
        cells = [c for c in r if c is not None]
        if len(cells) >= 2 and isinstance(cells[-1], (int, float)):
            label = str(cells[0]).strip()
            if not label or label.lower() == "account head":
                continue
            if label.lower() == "charges":
                summary["Total Charges"] = float(cells[-1])
            elif label in ("Realized P&L", "Unrealized P&L", "Other Credit & Debit"):
                summary[label] = float(cells[-1])
            else:
                charges[label] = float(cells[-1])

    other = _load_other_debits(path)
    return {"holdings": hold, "charges": charges, "summary": summary,
            "other": other, "period": period}


def _load_other_debits(path: str) -> pd.DataFrame:
    rows = _sheet_rows(path, "Other Debits and Credits")
    cols = ["particulars", "posting_date", "debit", "credit"]
    if rows is None:
        return pd.DataFrame(columns=cols)
    h = _find_header(rows, "particulars", "posting date")
    hdr = [str(c).strip().lower() if c else "" for c in rows[h]]
    ip, id_ = hdr.index("particulars"), hdr.index("posting date")
    idb = hdr.index("debit") if "debit" in hdr else None
    icr = hdr.index("credit") if "credit" in hdr else None
    rec = []
    for r in rows[h + 1:]:
        if r[ip] in (None, ""):
            continue
        rec.append((str(r[ip]).strip(), pd.to_datetime(r[id_]),
                    float(r[idb] or 0) if idb is not None else 0.0,
                    float(r[icr] or 0) if icr is not None else 0.0))
    return pd.DataFrame(rec, columns=cols)
