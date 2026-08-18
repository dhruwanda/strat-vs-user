"""Workbook for the implementation attribution."""
from __future__ import annotations
import pandas as pd
from openpyxl import Workbook
from .excel_report import (_title, _table, _auto_width, RS, RS2, PCT, PCT3, DATE,
                           S_FONT, N_FONT, WARN)

METHOD = [
    "MODEL QUANTITY: the reconstructed smallcase-prescribed quantity, validated "
    "against the investor's own unmodified orders.",
    "MODEL REFERENCE PRICE: the price at which the smallcase model/index carries "
    "the same transaction - rebalances at T1 (first trading day after the flagged "
    "date) OHLC average per documentation, T1 close indistinguishable in the "
    "independent index replication; invest events at the event day's close, since "
    "the index is an EOD close series. This is a model-construction price, not an "
    "execution price.",
    "The quantities SHOWN to the user are computed at order time from live prices "
    "(evidence sheet: 69/84 exact with execution VWAP + standard rounding vs 39/84 "
    "for the next basis). A Monday-morning order cannot embed Monday's OHLC.",
    "Implementation price effect = s x model_qty x (ref price - actual fill), "
    "s=+1 buy / -1 sell; positive = the investor transacted better than the model.",
    "Quantity deviation = actual qty - model qty; its cash component is valued at "
    "the actual fill.",
    "Per-leg identity, exact: actual value - model value = model_qty x (fill - "
    "ref) + (actual qty - model qty) x fill. No terminal price enters this table.",
    "QUANTITY SNAP RULE: a reconstructed-vs-executed deviation is adopted as the "
    "model's own quantity (no drift) when it is explainable without user "
    "modification - within +/-1 share, or reproducible with some price inside "
    "that day's traded low-high range, or the leg is the basket's cash balancer "
    "(its cash matches the other legs' residual; empirically the gold ETF plays "
    "this role), or an untraded model leg below the no-trade threshold. Only "
    "deviations none of these explain are classified 'user modification' and "
    "carried as quantity drift. Every leg's classification is shown.",
    "Dividends, costs and tax are separate layers and never explain this table.",
]


def build(out: dict, path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    impl = out["implementation"]

    ws = wb.create_sheet("Overview")
    r = _title(ws, "Model vs actual implementation",
               "Leg-level attribution first; P&L context, dividends, costs and tax "
               "follow separately.")
    r = _table(ws, out["overall"], r, fmts={"value": RS})
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and "Strategy return" in str(row[0].value):
            row[1].number_format = PCT
    ws.cell(row=r, column=1, value="Methodology").font = S_FONT
    r += 1
    for m in METHOD:
        c = ws.cell(row=r, column=1, value="- " + m)
        c.font = N_FONT
        r += 1
    _auto_width(ws)

    ws = wb.create_sheet("Implementation Effects")
    r = _title(ws, "Effects by side, event and stock",
               "Positive = investor better than the model reference.")
    r = _table(ws, impl["summary"], r, fmts={"amount": RS2})
    r = _title(ws, "Deviation classification (quantity snap rule)", row=r)
    r = _table(ws, impl["snap_counts"], r)
    r = _title(ws, "By event kind and side", row=r)
    r = _table(ws, impl["by_side"], r,
               fmts={"implementation_price_effect": RS,
                     "quantity_component": RS})
    r = _title(ws, "By event", row=r)
    r = _table(ws, impl["by_event"], r,
               fmts={"implementation_price_effect": RS,
                     "quantity_component": RS})
    r = _title(ws, "By stock", row=r)
    r = _table(ws, impl["by_stock"], r,
               fmts={"implementation_price_effect": RS,
                     "quantity_component": RS})
    _auto_width(ws)

    ws = wb.create_sheet("Attribution Lines")
    r = _title(ws, "Every model/actual leg pairing",
               "The table the eventual LLM layer consumes. BUY and SELL "
               "separately; deferred legs marked.")
    r = _table(ws, impl["lines"], r,
               fmts={"model_date": DATE, "actual_date": DATE, "model_qty": "#,##0",
                     "actual_qty": "#,##0", "qty_deviation": "#,##0",
                     "model_ref_price": RS2, "actual_price": RS2,
                     "implementation_price_effect": RS2, "price_component": RS2,
                     "quantity_component": RS2, "value_difference": RS2},
               wrap_cols=("status", "pairing"))
    _auto_width(ws)

    ws = wb.create_sheet("Rebalance Timeline")
    r = _title(ws, "Every rebalance: T0, T1, conventions, application",
               "T0 = flagged model date = start of the new version. The index "
               "transitions on T1. The investor may act before or after T1; "
               "negative lag means the order went in on T0 itself, before the "
               "index's own transition day.")
    r = _table(ws, out["timeline"], r,
               fmts={"T0_model_rebalance_date": DATE, "version_effective_from": DATE,
                     "T1_first_trading_day": DATE, "investor_applied_on": DATE,
                     "lag_trading_days": "0"},
               wrap_cols=("index_quantity_price_basis",
                          "user_quantity_price_basis", "rounding_rule"))
    _auto_width(ws)

    ws = wb.create_sheet("Index Replication")
    r = _title(ws, "Official index vs independent reconstruction",
               "The index is rebuilt from constituents, weights and EOD prices "
               "under each candidate convention; the anchor scales once at the "
               "first transition, everything after is out-of-sample.")
    r = _table(ws, out["replication_summary"], r,
               fmts={"mean_abs_diff_pct": PCT3, "max_abs_diff_pct": PCT3,
                     "worst_date": DATE})
    best = out["replication_summary"].sort_values("mean_abs_diff_pct").iloc[0]
    d = out["replication_daily"]
    d = d[d["convention"] == best["convention"]]
    ws.cell(row=r, column=1,
            value=f"Daily series, best convention ({best['convention']}):").font = S_FONT
    r = _table(ws, d[["date", "segment", "official", "replica", "diff_pct", "note"]],
               r + 1, fmts={"date": DATE, "official": "0.00", "replica": "0.00",
                            "diff_pct": PCT3})
    _auto_width(ws)

    ws = wb.create_sheet("Qty Convention Evidence")
    r = _title(ws, "Which price basis and rounding reproduce the quantities?",
               "Tested against every observed quantity. Invest events answer "
               "decisively: order-time live prices with standard rounding. "
               "Rebalances are consistent with the same but blur within +/-1 "
               "share; floor vs round is inconclusive there and flagged as such.")
    ws.cell(row=r, column=1, value="Invest events").font = S_FONT
    r = _table(ws, out["evidence_invest"], r + 1)
    ws.cell(row=r, column=1, value="Rebalances").font = S_FONT
    r = _table(ws, out["evidence_rebalance"], r + 1)
    _auto_width(ws)

    ws = wb.create_sheet("Deferred Legs")
    r = _title(ws, "Standalone trades attached as deferred event legs",
               "A leg can fail in the basket (circuit limit) and be repaired over "
               "the following days. Attachment requires a model shortfall in that "
               "symbol and direction, the configured window, and a quantity within "
               "the shortfall. None were needed for this dataset; the mechanism is "
               "unit-tested.")
    r = _table(ws, out["deferred_repairs"], r,
               fmts={"trade_date": DATE, "qty": "#,##0", "shortfall": "#,##0"})
    _auto_width(ws)

    ws = wb.create_sheet("Dividends")
    r = _title(ws, "Dividend statement, attributed to the smallcase",
               "A dividend is attributed only to the extent the smallcase held "
               "the shares before the ex-date, pro-rata. Dividends on personal "
               "or pre-existing holdings attribute to zero. Kept out of the "
               "implementation attribution by design.")
    if len(out.get("dividend_statement", [])):
        r = _table(ws, out["dividend_statement"], r,
                   fmts={"ex_date": DATE, "qty": "#,##0", "dps": RS2,
                         "amount": RS2, "smallcase_qty_on_ex_date": "#,##0",
                         "attributed_amount": RS2},
                   wrap_cols=("attribution_note",))
        ds = out["dividend_statement"]
        ws.cell(row=r, column=1,
                value=f"Statement total {ds['amount'].sum():,.2f}; attributed to "
                      f"the smallcase {ds['attributed_amount'].sum():,.2f}. The "
                      "smallcase page shows its own figure; any difference is a "
                      "page-vs-statement timing or inclusion difference and is "
                      "shown, not forced.").font = N_FONT
    else:
        ws.cell(row=r, column=1,
                value="No dividend statement supplied; the configured figure "
                      "from the smallcase page is used and labelled as such."
                ).font = N_FONT
    _auto_width(ws)

    ws = wb.create_sheet("Tax Base")
    r = _title(ws, "Deterministic tax base - no rates applied",
               "The engine asserts realised gains and losses by financial year, "
               "asset class and holding-period term. Applicable treatment and "
               "rates are for the interpretation layer or a tax adviser; the "
               "engine refuses to guess them.")
    r = _table(ws, out.get("tax_base", pd.DataFrame()), r,
               fmts={"realised_gain": RS2, "realised_loss": RS2, "net": RS2,
                     "lots": "0"})
    _auto_width(ws)

    ws = wb.create_sheet("Missing Prices")
    r = _title(ws, "Every price the data source could not supply",
               "Excluded and listed, never substituted.")
    r = _table(ws, out["missing_prices"], r, fmts={"date": DATE},
               wrap_cols=("purpose",))
    if len(out["splits_detected"]):
        ws.cell(row=r, column=1,
                value="Corporate actions detected in the span:").font = S_FONT
        r = _table(ws, out["splits_detected"], r + 1, fmts={"date": DATE})
    _auto_width(ws)

    wb.save(path)
    return path
