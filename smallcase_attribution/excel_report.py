"""
Render the pipeline output to an Excel workbook.

Convention: detail sheets carry the pipeline's computed values; summary sheets
are built from formulas that reference those detail sheets, so the workbook
recalculates if a detail figure is corrected by hand.
"""
from __future__ import annotations
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
T_FONT = Font(name=FONT, size=14, bold=True, color="1F3864")
S_FONT = Font(name=FONT, size=10, bold=True, color="1F3864")
B_FONT = Font(name=FONT, size=10)
N_FONT = Font(name=FONT, size=9, italic=True, color="595959")
LINK = Font(name=FONT, size=10, color="0000FF", underline="single")
GREEN = Font(name=FONT, size=10, color="008000")
WARN = PatternFill("solid", fgColor="FFF2CC")
BAND = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))

RS = '#,##0;(#,##0);"-"'
RS2 = '#,##0.00;(#,##0.00);"-"'
PCT = '0.00%;(0.00%);"-"'
PCT3 = '0.000%;(0.000%);"-"'
DATE = "yyyy-mm-dd"


def _auto_width(ws, max_w=52):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(10, w + 2), max_w)


def _title(ws, text, note=None, row=1):
    ws.cell(row=row, column=1, value=text).font = T_FONT
    if note:
        ws.cell(row=row + 1, column=1, value=note).font = N_FONT
        return row + 3
    return row + 2


def _table(ws, df: pd.DataFrame, start_row: int, fmts: dict | None = None,
           wrap_cols=(), band=True) -> int:
    """Write a frame with a styled header. Returns the row after the table."""
    fmts = fmts or {}
    if df is None or not len(df):
        ws.cell(row=start_row, column=1, value="(no rows)").font = N_FONT
        return start_row + 2
    for j, c in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=str(c).replace("_", " "))
        cell.font, cell.fill = H_FONT, H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (_, r) in enumerate(df.iterrows(), start=1):
        for j, c in enumerate(df.columns, start=1):
            v = r[c]
            if isinstance(v, (list, set, tuple)):
                v = ", ".join(map(str, sorted(v)))
            if isinstance(v, (pd.Timestamp,)):
                v = v.to_pydatetime()
            elif isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating,)):
                v = None if np.isnan(v) else float(v)
            elif isinstance(v, (np.bool_,)):
                v = bool(v)
            cell = ws.cell(row=start_row + i, column=j, value=v)
            cell.font = B_FONT
            cell.border = THIN
            if band and i % 2 == 0:
                cell.fill = BAND
            if c in fmts:
                cell.number_format = fmts[c]
            elif isinstance(v, float):
                cell.number_format = RS2
            if c in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(df) + 3


def build(res: dict, path: str, reported: dict | None = None,
          source_note: str = "") -> str:
    reported = reported or {}
    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------
    def sheet(name):
        return wb.create_sheet(name[:31])

    # 1. Read me -------------------------------------------------------
    ws = sheet("Read Me First")
    r = _title(ws, "smallcase: strategy vs implementation vs net outcome",
               "Three questions kept apart on purpose. Read the layers in order.")
    guide = pd.DataFrame([
        ("Summary", "The three layers side by side. Start here."),
        ("Gap Bridge", "Why the actual result differs from the strategy. Read the "
                       "status column: only rows marked 'measured' carry a number."),
        ("Strategy Index", "Layer A. smallcase's official index rebased to the first "
                           "investment date."),
        ("Rebalance Timing", "Model rebalance dates vs the date each was actually applied."),
        ("Events", "Every smallcase order detected, and how well the recommended "
                   "quantities could be reproduced."),
        ("Recommended Qty", "Line-by-line reconstruction: what smallcase would have "
                            "recommended vs what was traded."),
        ("Trades Attributed", "Every broker trade with an explicit in/out decision and "
                              "the reason for it."),
        ("Positions", "Current smallcase holdings, average cost, market value, and drift "
                      "from prescribed weights."),
        ("Weight Drift", "Current weights against the live model version."),
        ("Realised P&L", "Booked gains on the smallcase's average-cost basis."),
        ("Reconciliation", "Reconstruction against the figures smallcase reports."),
        ("Cost Breakdown", "What the smallcase cost to run, and what could not be attributed."),
        ("Cost Calibration", "How per-trade charges were derived and checked."),
        ("Tax Estimate", "Estimated capital-gains impact. An estimate, not a liability."),
        ("Tax Lots FIFO", "The FIFO lot matching behind the tax estimate."),
        ("Stock Attribution", "Per-stock contribution to the result."),
        ("Assumptions", "Every judgement made, with the evidence for it."),
        ("Limitations", "What could not be established, and the data that would fix it."),
        ("Mapping", "Constituent name to broker symbol, with confidence."),
        ("Constituents", "Normalised model composition history."),
    ], columns=["sheet", "what it contains"])
    r = _table(ws, guide, r, wrap_cols=("what it contains",))
    ws.cell(row=r, column=1, value="Sources").font = S_FONT
    r += 1
    for line in (source_note or "").split("\n"):
        if line.strip():
            ws.cell(row=r, column=1, value=line).font = N_FONT
            r += 1
    _auto_width(ws)

    # 2. Summary -------------------------------------------------------
    ws = sheet("Summary")
    lay = res["layers"]
    r = _title(ws, "Strategy vs implementation vs net outcome",
               "Layer A is the model. A' puts the model on the investor's own cash-flow "
               "timing, which is the only like-for-like comparison. B is what actually "
               "happened. C is what was kept.")
    r = _table(ws, lay, r, fmts={"value_rs": RS, "return_pct": PCT},
               wrap_cols=("question", "measure", "basis"))
    impl = res["implementation"]
    st = res["strategy_summary"]
    ws.cell(row=r, column=1, value="Key figures").font = S_FONT
    r += 1
    key = pd.DataFrame([
        ("Money put in", impl["money_put_in"], RS),
        ("Current investment (cost of what is still held)", impl["current_investment"], RS),
        ("Current market value", impl["current_value"], RS),
        ("Unrealised return", impl["current_returns"], RS),
        ("Realised return", impl["realized_returns"], RS),
        ("Dividends", impl["dividends"], RS),
        ("Total return", impl["total_returns"], RS),
        ("Total return %", impl["total_return_pct"], PCT),
        ("Strategy index return over the same window", st["absolute_return"], PCT),
        ("Strategy index max drawdown", st["max_drawdown"], PCT),
        ("Attributable costs", -sum(res["cost_summary"]["smallcase"]), RS),
        ("Estimated tax", 0.0 if res["tax_estimate"].empty else
         -float(res["tax_estimate"].loc[res["tax_estimate"]["view"].str.startswith("marginal"),
                                        "estimated_tax"].sum()), RS),
    ], columns=["metric", "value", "fmt"])
    for j, c in enumerate(["metric", "value"], start=1):
        cell = ws.cell(row=r, column=j, value=c)
        cell.font, cell.fill = H_FONT, H_FILL
    for i, (_, x) in enumerate(key.iterrows(), start=1):
        ws.cell(row=r + i, column=1, value=x["metric"]).font = B_FONT
        c = ws.cell(row=r + i, column=2,
                    value=(None if x["value"] is None or
                           (isinstance(x["value"], float) and np.isnan(x["value"]))
                           else float(x["value"])))
        c.font, c.number_format, c.border = B_FONT, x["fmt"], THIN
    r += len(key) + 2
    ws.cell(row=r, column=1,
            value="Layer C is labelled Net Return After Costs & Taxes. It is not an "
                  "after-tax certainty: see the Tax Estimate sheet.").font = N_FONT
    _auto_width(ws)

    # 3. Gap bridge ----------------------------------------------------
    ws = sheet("Gap Bridge")
    r = _title(ws, "Why the actual result differs from the strategy",
               "Rows marked 'measured' are computed from the data. Rows marked "
               "'exposure measured' quantify how much of a driver was present, not how "
               "much money it made or cost. Nothing is allocated to a driver without "
               "evidence.")
    r = _table(ws, res["gap_bridge"], r, fmts={"amount": RS},
               wrap_cols=("item", "evidence", "status"))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and "UNEXPLAINED" in str(row[0].value):
            for c in row:
                c.fill = WARN
    _auto_width(ws)

    # 4. Strategy index -------------------------------------------------
    ws = sheet("Strategy Index")
    r = _title(ws, "Layer A: smallcase official index, rebased to the first investment",
               "Price return only. No dividends, no transaction costs, no taxes, and no "
               "money actually deployed. Source: smallcase timeline export.")
    summ = pd.DataFrame([{k: v for k, v in res["strategy_summary"].items()}])
    r = _table(ws, summ, r, fmts={"absolute_return": PCT, "annualised_return": PCT,
                                  "max_drawdown": PCT, "start_date": DATE,
                                  "end_date": DATE, "normalised_end": "0.00",
                                  "base_index_value": "0.00",
                                  "end_index_value": "0.00"})
    idx = res["strategy_index"][["date", "index_value", "normalised_index",
                                 "cumulative_return", "rebalance_flag"]]
    r = _table(ws, idx, r, fmts={"date": DATE, "index_value": "0.00",
                                 "normalised_index": "0.00",
                                 "cumulative_return": PCT})
    _auto_width(ws)

    # 5. Rebalance timing ------------------------------------------------
    ws = sheet("Rebalance Timing")
    r = _title(ws, "Model rebalance dates vs the dates they were applied",
               "The index switches constituents on the model date. A later application "
               "leaves the investor holding the previous book for the lag window. The "
               "index move during the lag is shown as evidence of exposure; it is not "
               "the profit or loss caused, which needs per-constituent prices.")
    r = _table(ws, res["timing_lag"], r,
               fmts={"model_rebalance_date": DATE, "investor_applied_on": DATE,
                     "index_at_model": "0.00", "index_at_applied": "0.00",
                     "index_move_during_lag": PCT, "lag_days": "0"})
    r = _table(ws, res["rebalance_calendar"], r,
               fmts={"model_rebalance_date": DATE, "investor_applied_on": DATE,
                     "index_value_on_date": "0.00", "lag_days": "0"})
    _auto_width(ws)

    # 6. Events -----------------------------------------------------------
    ws = sheet("Events")
    r = _title(ws, "smallcase orders detected in the tradebook",
               "An event is a basket: several symbols filled at one execution timestamp. "
               "A transaction fee on the same date confirms fresh money; rebalances carry "
               "no fee. Value deviation is the honest measure of how well the recommended "
               "quantities were reproduced.")
    cols = ["event_id", "date", "kind", "version", "n_symbols", "buy_value", "sell_value",
            "net_cash", "smallcase_fee_on_date", "model_rebalance_date", "lag_days",
            "reconstruction_basis", "lines_reconstructible", "lines_consistent",
            "lines_not_reconstructible", "abs_value_deviation",
            "value_deviation_pct_of_event", "portfolio_value_basis", "portfolio_value",
            "probe_dispersion_cv", "round_amount_tested", "round_amount_consistent",
            "unpriced_holdings"]
    ev = res["event_summary"][[c for c in cols if c in res["event_summary"]]]
    r = _table(ws, ev, r,
               fmts={"date": DATE, "model_rebalance_date": DATE, "buy_value": RS,
                     "sell_value": RS, "net_cash": RS, "abs_value_deviation": RS2,
                     "value_deviation_pct_of_event": PCT3, "portfolio_value": RS,
                     "probe_dispersion_cv": PCT, "round_amount_tested": RS,
                     "lag_days": "0"},
               wrap_cols=("reconstruction_basis", "portfolio_value_basis",
                          "unpriced_holdings"))
    ws.cell(row=r, column=1,
            value="portfolio_value on rebalance rows is inferred from the traded legs "
                  "themselves. It measures weight fidelity; it is NOT a reliable "
                  "independent valuation of the portfolio on that date.").font = N_FONT
    _auto_width(ws)

    # 7. Recommended quantities -------------------------------------------
    ws = sheet("Recommended Qty")
    r = _title(ws, "Reconstructed recommendation vs what was actually traded",
               "Investment events: qty = round(((value before + amount) x weight - held x "
               "price) / price). Rebalance events: each touched constituent driven to "
               "weight x portfolio value. Exact share equality is not achievable because "
               "smallcase sizes the order off a live quote the tradebook does not record.")
    r = _table(ws, res["recommended_lines"], r,
               fmts={"event_ts": "yyyy-mm-dd hh:mm:ss", "weight": PCT,
                     "price": RS2, "value_diff": RS2, "weight_actual": PCT},
               wrap_cols=("status",))
    _auto_width(ws)

    # 8. Trades ------------------------------------------------------------
    ws = sheet("Trades Attributed")
    r = _title(ws, "Every broker trade, in or out of the smallcase, with the reason",
               "Rule: a trade belongs to the smallcase only if it filled inside a basket "
               "of several symbols sharing one execution timestamp. Standalone orders in "
               "the same stock are excluded and listed here so the decision is auditable.")
    t = res["per_trade_costs"]
    keep = ["exec_ts", "trade_date", "symbol", "exchange", "trade_type", "quantity",
            "price", "value", "attribution", "event_id", "event_kind", "reason", "flag",
            "stt", "exchange_txn", "sebi_turnover", "ipft", "stamp_duty", "brokerage",
            "gst", "total_charges"]
    r = _table(ws, t[[c for c in keep if c in t]], r,
               fmts={"exec_ts": "yyyy-mm-dd hh:mm:ss", "trade_date": DATE,
                     "price": RS2, "value": RS2, "quantity": "#,##0"},
               wrap_cols=("reason",))
    _auto_width(ws)

    # 9. Positions ----------------------------------------------------------
    ws = sheet("Positions")
    r = _title(ws, "Current smallcase holdings",
               "Average-cost basis, which is the convention smallcase reports on. "
               "Quantities exclude any units of the same scrip bought outside the "
               "smallcase.")
    r = _table(ws, res["positions"], r,
               fmts={"qty": "#,##0", "cost_basis": RS, "avg_price": RS2,
                     "current_price": RS2, "market_value": RS, "unrealized_pnl": RS,
                     "weight_actual": PCT, "weight_prescribed": PCT,
                     "weight_deviation": PCT})
    _auto_width(ws)

    ws = sheet("Weight Drift")
    r = _title(ws, "Current weights against the live model version",
               "The index resets to exact prescribed weights at every rebalance. A real "
               "book does not, because smallcase minimises churn. This drift is one "
               "candidate driver of the gap on the Gap Bridge sheet.")
    r = _table(ws, res["weight_drift"], r,
               fmts={"market_value": RS, "weight_actual": PCT,
                     "weight_prescribed": PCT, "weight_deviation": PCT,
                     "deviation_bps": "#,##0", "value_vs_prescribed": RS})
    _auto_width(ws)

    # 10. Realisations --------------------------------------------------------
    ws = sheet("Realised P&L")
    r = _title(ws, "Booked gains and losses, smallcase average-cost basis",
               "This reproduces the Realized Returns figure on the smallcase Investments "
               "page. It is NOT the tax basis: see Tax Lots FIFO.")
    r = _table(ws, res["realisations"], r,
               fmts={"exec_ts": "yyyy-mm-dd hh:mm:ss", "date": DATE, "qty": "#,##0",
                     "sell_price": RS2, "avg_cost": RS2, "realized_pnl": RS2,
                     "sell_value": RS2, "cost_released": RS2})
    _auto_width(ws)

    # 11. Reconciliation ------------------------------------------------------
    ws = sheet("Reconciliation")
    r = _title(ws, "Reconstruction against smallcase's own reported figures",
               "Nothing here is forced to match. Where a difference remains it is named, "
               "not absorbed.")
    r = _table(ws, res["reconciliation"], r,
               fmts={"smallcase_reported": RS, "reconstructed": RS,
                     "difference": RS2, "difference_pct": PCT3},
               wrap_cols=("definition_used", "likely_explanation", "data_required"))
    _auto_width(ws)

    # 12/13. Costs -------------------------------------------------------------
    ws = sheet("Cost Breakdown")
    r = _title(ws, "What the smallcase cost to run",
               "Charges are derived per trade from a rate card and then scaled so each "
               "head sums exactly to the broker's reported account total. Only the "
               "smallcase column is attributed to this strategy.")
    cs = res["cost_summary"]
    r0 = r
    r = _table(ws, cs, r, fmts={"smallcase": RS2, "outside": RS2, "total": RS2},
               wrap_cols=("basis",))
    tot = r - 2
    ws.cell(row=r, column=1, value="TOTAL").font = S_FONT
    for j, col in enumerate(["smallcase", "outside", "total"], start=2):
        letter = get_column_letter(j)
        c = ws.cell(row=r, column=j, value=f"=SUM({letter}{r0+1}:{letter}{tot})")
        c.font, c.number_format = S_FONT, RS2
    r += 2
    ws.cell(row=r, column=1, value="Costs identified but NOT attributable").font = S_FONT
    r = _table(ws, res["unattributed_costs"], r + 1, fmts={"amount": RS2},
               wrap_cols=("reason",))
    _auto_width(ws)

    ws = sheet("Cost Calibration")
    r = _title(ws, "How each charge head was derived and checked",
               "A calibration factor near 1.0 means the published rate reproduces the "
               "broker's total. A factor far from 1.0 flags a rate assumption that does "
               "not hold; the attributed amount is still exact at account level because "
               "the head is scaled to the reported total.")
    r = _table(ws, res["cost_calibration"], r,
               fmts={"broker_reported_total": RS2, "derived_before_calibration": RS2,
                     "calibration_factor": "0.000"}, wrap_cols=("method",))
    _auto_width(ws)

    # 14. Tax --------------------------------------------------------------------
    ws = sheet("Tax Estimate")
    r = _title(ws, "Estimated capital-gains impact of the smallcase's realised trades",
               "An estimate, not a liability. The exemption and loss set-off are annual "
               "and taxpayer-level, so the true figure depends on the whole portfolio.")
    r = _table(ws, res["tax_estimate"], r,
               fmts={c: RS2 for c in ("short_term_gain", "long_term_gain",
                                      "unclassified_gain", "short_term_after_setoff",
                                      "long_term_after_setoff", "ltcg_exemption_applied",
                                      "stcg_tax", "ltcg_tax", "cess", "estimated_tax",
                                      "loss_carried_forward")},
               wrap_cols=("view",))
    ws.cell(row=r, column=1, value="Notes").font = S_FONT
    r += 1
    for n in res["tax_notes"]:
        ws.cell(row=r, column=1, value="- " + n).font = N_FONT
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    if len(res["intraday_legs"]):
        r += 1
        ws.cell(row=r, column=1,
                value="Same-day offsetting legs, treated as intraday rather than "
                      "capital gains").font = S_FONT
        r = _table(ws, res["intraday_legs"], r + 1,
                   fmts={"date": DATE, "qty": "#,##0", "buy_price": RS2,
                         "sell_price": RS2, "pnl": RS2, "turnover": RS2})
    _auto_width(ws)

    ws = sheet("Tax Lots FIFO")
    r = _title(ws, "FIFO lot matching behind the tax estimate",
               "FIFO runs across every unit of each scrip, because a demat holding is "
               "fungible. Where 'lot from smallcase' is FALSE, a smallcase sale consumed "
               "units the investor had bought independently.")
    r = _table(ws, res["fifo_realisations"], r,
               fmts={"buy_date": DATE, "sell_date": DATE, "qty": "#,##0",
                     "buy_price": RS2, "sell_price": RS2, "cost": RS2,
                     "proceeds": RS2, "gain": RS2, "holding_days": "#,##0"})
    _auto_width(ws)

    # 15. Stock attribution -------------------------------------------------------
    ws = sheet("Stock Attribution")
    r = _title(ws, "Per-stock contribution",
               "Covers every symbol the smallcase held at any point in the period, "
               "including those fully exited.")
    r = _table(ws, res["stock_attribution"], r,
               fmts={"qty_bought": "#,##0", "qty_sold": "#,##0", "qty_held": "#,##0",
                     "buy_value": RS, "sell_value": RS, "avg_cost": RS2,
                     "cost_basis": RS, "current_price": RS2, "market_value": RS,
                     "realized_pnl": RS, "unrealized_pnl": RS, "total_pnl": RS,
                     "attributable_trading_costs": RS2,
                     "contribution_to_realized_pnl_pct": PCT,
                     "contribution_to_total_pnl_pct": PCT})
    _auto_width(ws)

    # 16/17. Assumptions and limitations --------------------------------------------
    ws = sheet("Assumptions")
    r = _title(ws, "Every judgement made, and the evidence for it")
    r = _table(ws, res["assumptions"], r, wrap_cols=("assumption", "evidence"))
    _auto_width(ws)

    ws = sheet("Limitations")
    r = _title(ws, "What could not be established from the supplied files",
               "Each row names the exact extra input that would resolve it.")
    r = _table(ws, res["limitations"], r,
               wrap_cols=("limitation", "impact", "data_required"))
    _auto_width(ws)

    # 18/19. Mapping and constituents -------------------------------------------------
    ws = sheet("Mapping")
    r = _title(ws, "Constituent name to broker symbol",
               "Resolved from basket membership and rebalance additions/removals, with "
               "name similarity only as a tiebreak. Rows marked 'review' were won by a "
               "small margin and should be confirmed before the numbers are relied on.")
    r = _table(ws, res["mapping_diagnostics"], r,
               fmts={"score": "0.000", "margin": "0.000", "name_similarity": "0.000"})
    _auto_width(ws)

    ws = sheet("Constituents")
    r = _title(ws, "Normalised model composition history",
               "Forward-filled from the smallcase timeline export, where the date range "
               "appears only on the first row of each version block.")
    r = _table(ws, res["constituents"], r,
               fmts={"version_start": DATE, "version_end": DATE, "weight": PCT})
    _auto_width(ws)

    wb.save(path)
    return path
